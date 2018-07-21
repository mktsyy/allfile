# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
from django.http import HttpResponse
from fill.models import Person
from fill.models import Information
import json
from django.http import JsonResponse
from PIL import Image
import os
import othervarible
# Create your views here.
# 
import sys
defaultencoding = 'utf-8'
if sys.getdefaultencoding() != defaultencoding:
    reload(sys)
    sys.setdefaultencoding(defaultencoding)



def fill(request):
	a = request.GET.get('a',0)
	b = request.GET.get('b',0)
	c = request.GET.get('c',0)
	d = request.GET.get('d',0)
	e = request.GET.get('e',0)
	f = request.GET.get('f',0)
	g = request.GET.get('g',0)
	h = request.GET.get('h',0)
	z = request.GET.get('z',0)


	if a:
		#写入数据库
		Person.objects.create(tier=a,houseinfo=b,housecf=c,rentfl=d,housecf5=e)
# 		returndata = Person.objects.all()[len(Person.objects.all())-1]

# 		readscriptpath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))+"\chrome_fill_login\\readscript.js"
# 		readscriptpath.replace("\\","\\\\")
# 		with open(readscriptpath,"w+") as f:
# 			f.write('console.log("'+'%c'+'%s", "color: #e40c0c;font-size: xx-large;")\n' % (returndata.tier).encode("utf-8"))
# 			f.write('document.getElementById("muban").click();\n')
# 			f.write('document.getElementById("blockshowname").value="%s"\n' % (returndata.tier).encode("utf-8"))
# 			f.write('document.getElementsByName("renttype")[1].checked = true\n')
# 			f.write('document.getElementById("price").value=%s\n' % returndata.rentfl.encode("utf-8"))
# 			f.write('document.getElementById("buildarea").value=%s\n' % returndata.housecf.encode("utf-8").split('㎡')[0])
# 			f.write('document.getElementsByName("room")[0].value=%s\n' % returndata.houseinfo.encode("utf-8").split("室")[0])
# 			f.write('document.getElementsByName("hall")[0].value=%s\n' % returndata.houseinfo.encode("utf-8").split("室")[1].split("厅")[0])
# 			f.write('document.getElementsByName("toilet")[0].value=%s\n' % returndata.houseinfo.encode("utf-8").split("室")[1].split("厅")[1][0])

# 			if int(returndata.housecf5.encode("utf-8").split("/")[0])==0:
# 				f.write('document.getElementsByName("floor")[0].value=%s\n' % "3")
# 			else:
# 				f.write('document.getElementsByName("floor")[0].value=%s\n' % returndata.housecf5.encode("utf-8").split("/")[0])

# 			if int(returndata.housecf5.encode("utf-8").split("/")[1].split("层")[0])==0:
# 				f.write('document.getElementsByName("totalfloor")[0].value=%s\n' % "7")
# 			else:
# 				f.write('document.getElementsByName("totalfloor")[0].value=%s\n' % returndata.housecf5.encode("utf-8").split("/")[1].split("层")[0])
			

# 			kkkk = '''document.getElementsByName("fitment")[0].value="精装"
# document.getElementsByName("forward")[0].value="南北通透"
# document.getElementsByName("payment")[0].value="月付"

# document.getElementById("jsAllcheckbox").checked=true

# for (var i = 10 - 1; i >= 0; i--) {
# 	document.getElementsByName("equipment[]")[i].checked=true
# };\n
# '''

# 			f.write(kkkk)

# 			#主次卧调换
# 			if int(returndata.rentfl)%100 == 0:
# 				f.write('document.getElementsByName("rentroom")[0].value="次卧"\n')
# 			else:
# 				f.write('document.getElementsByName("rentroom")[0].value="主卧"\n')

# 			#标题调换
# 			title = (returndata.tier).encode("utf-8")+"环境优雅 闹中取静 交通便利 温馨舒适 阳光充足"
# 			if int(returndata.rentfl)%100 == 0:
# 				f.write('document.getElementById("address").value="%s"\n' % title)
# 			elif int(returndata.rentfl)%50 == 0:
# 				title = (returndata.tier).encode("utf-8")+"采光好 空间大 小区绿化好 配套齐全 环境舒适"
# 				f.write('document.getElementById("address").value="%s"\n' % title)
# 			else:
# 				title = (returndata.tier).encode("utf-8")+"环境整洁清爽 采光格局好 冬暖夏凉 交通便利"
# 				f.write('document.getElementById("address").value="%s"\n' % title)

			
# 			selectpic = '''
# setTimeout(function(){
# for (var i = 1; i < 5; i++) {
# 	document.getElementsByClassName("btn")[i].click()
# 	}
# },1000)
# ''' 
# 			f.write(selectpic)
# 			#拿取小区地址自动填入
# 			idjavascript = '''
# function myTimer() {
#     var xmlhttp;
#     xmlhttp = new XMLHttpRequest();
#     xmlhttp.onreadystatechange = function() {
#         if (xmlhttp.readyState == 4 && xmlhttp.status == 200) {
            
#             console.log(JSON.parse(xmlhttp.responseText))
#             document.getElementById("district").value = JSON.parse(xmlhttp.responseText).district
#             document.getElementById("street").value = JSON.parse(xmlhttp.responseText).streetname
#             document.getElementById("blockaddress").value = JSON.parse(xmlhttp.responseText).address
#             document.getElementById("blockshowname").value=JSON.parse(xmlhttp.responseText).blockname
#             window.scroll(100,1650);
#         } 
#     }
#     xmlhttp.open("GET", "http://nj.zsb.house365.com/ajax/find_block_detail/?id=%s" , true);
    
#     xmlhttp.send();

# };
# myTimer();\n

# document.onkeydown=function(event){
#       var e = event || window.event || arguments.callee.caller.arguments[0];
                
#        // if(e && e.keyCode==13){ // enter 键
#        if(e && e.keyCode==32){ // space 键
#            document.getElementById("jsBtnSubmit_rent").click();
#       }
#   }; 

# 			''' % (str(getid((returndata.tier).encode("utf-8"))))

# 			f.write(idjavascript)
			
# 			# content = '''
# 			# 嗨住的优质房源
# 			# 【便利的交通&成熟的配套】
# 			# 超市，银行，医院，学校，便利店，菜市场，购物广场应有尽有,安全性高、位置安静、靠近花园、采光什么都非常好
# 			# 【装修精致\配套成熟的整租房源】

# 			# 厨房、卫生间、热水器、空调、WIFI,设备齐全，拎包就能入住,还有送车位的哦~
# 			# 【有趣善良的合租室友】
# 			# 那些直播里的萌妹纸、那些陆家嘴的金融GG、那些一身正气的程序员，可能就是你隔壁的室友~那些合租的小伙伴都是素质住户
# 			# 【品质小区】
# 			# 小区都处于繁华地段、房型正、佳、环境非常优美、都是一些次新小区
# 			# 嗨住的优质服务
# 			# * 无任何中介费
# 			# 嗨住房源实地勘察，保证嗨住租房平台房源可租信息实时可靠
# 			# * 房源广（7x24不断更新中）
# 			# 上海房源48396、北京16411、杭州23495、南京12055
# 			# * 直呼房东
# 			# 租客可通过嗨住租房上公开的房东电话，直接呼叫联系房东，询问房屋信息、预约看 房、达成交易。
# 			# * 免费预约
# 			# 租客可直接在嗨住租房平台提交对某套房的看房预约请求，嗨住顾问将免费为租客协调与房东的见面时间，与房东直接见面，达成交易。
# 			# * 租后服务
# 			# 嗨住还为租客提供了房租分期、室友征选、搬家维修等系列租后服务。
# 			# 更多房源欢迎致电4008180555(上海4008180555，杭州4008170019，南京4001606145）
# 			# 或
# 			# 至任意手机应用市场下载“嗨住租房”app


# 			# '''

			

		return HttpResponse (a)
	if f:
		rdict = {}
		returndata = Person.objects.all()[len(Person.objects.all())-1]
		rdict['tier'] = returndata.tier
		rdict['houseinfo'] = returndata.houseinfo
		rdict['housecf'] = returndata.housecf
		rdict['rentfl'] = returndata.rentfl
		rdict['housecf5'] = returndata.housecf5
		


		return HttpResponse(json.dumps(rdict),content_type="application/json")
		# return HttpResponse("here")
	if g:
		# print eval(g)['address']
		Information.objects.create(address=eval(g)['address'],streetname=eval(g)['streetname'],district=eval(g)['district'],blockid=eval(g)['blockid'],blockname=eval(g)['blockname'])

	if h:
		Person.objects.all().delete();

	if z:
		
		return render(request,"23.html") 
	
		
	k = ""
	data = Person.objects.all()
	for i in data:
		k+=i.tier+i.houseinfo+i.housecf+i.rentfl+i.housecf5+"---"
	return HttpResponse(k)


	return HttpResponse (a)



def getid(a):

	idpath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))+"\id.html"
	idpath.replace("\\","\\\\")
	with open(idpath,"r") as f:
		for i in f.readlines():
			if a in i:
				id =  i.split('addlishi(')[1].split(')')[0]
				# print i
				break
	return id

def temptestselect(request):
	testselect = request.GET.get("test")
	# print str(testselect)
	print(os.getcwd())
	return HttpResponse (str(testselect))


def writedata(request,datanum):
	returndata = Person.objects.all()[int(datanum)]
	print returndata.tier
	housenum = len(Person.objects.all())
	returntier = {"tier":(returndata.tier).encode("utf-8"),"price":returndata.rentfl.encode("utf-8"),"housenum":str(housenum)}

	readscriptpath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))+"\chrome_fill_login\\readscript.js"
	readscriptpath.replace("\\","\\\\")
	with open(readscriptpath,"w+") as f:
		f.write('console.log("'+'%c'+'%s", "color: #e40c0c;font-size: xx-large;")\n' % (returndata.tier).encode("utf-8"))
		f.write('document.getElementById("muban").click();\n')
		f.write('document.getElementById("blockshowname").value="%s"\n' % (returndata.tier).encode("utf-8"))
		f.write('document.getElementsByName("renttype")[1].checked = true\n')
		f.write('document.getElementById("price").value=%s\n' % returndata.rentfl.encode("utf-8"))
		f.write('document.getElementById("buildarea").value=%s\n' % returndata.housecf.encode("utf-8").split('m²')[0])
		f.write('document.getElementsByName("room")[0].value=%s\n' % returndata.houseinfo.encode("utf-8").split("室")[0])
		f.write('document.getElementsByName("hall")[0].value=%s\n' % returndata.houseinfo.encode("utf-8").split("室")[1].split("厅")[0])
		f.write('document.getElementsByName("toilet")[0].value=%s\n' % returndata.houseinfo.encode("utf-8").split("室")[1].split("厅")[1][0])

		if int(returndata.housecf5.encode("utf-8").split("/")[0])==0:
			f.write('document.getElementsByName("floor")[0].value=%s\n' % "3")
		else:
			f.write('document.getElementsByName("floor")[0].value=%s\n' % returndata.housecf5.encode("utf-8").split("/")[0])

		if int(returndata.housecf5.encode("utf-8").split("/")[1].split("层")[0])==0:
			f.write('document.getElementsByName("totalfloor")[0].value=%s\n' % "7")
		else:
			f.write('document.getElementsByName("totalfloor")[0].value=%s\n' % returndata.housecf5.encode("utf-8").split("/")[1].split("层")[0])
		

		kkkk = '''document.getElementsByName("fitment")[0].value="精装"
document.getElementsByName("forward")[0].value="南北通透"
document.getElementsByName("payment")[0].value="月付"
document.getElementById("jsAllcheckbox").checked=true

for (var i = 10 - 1; i >= 0; i--) {
document.getElementsByName("equipment[]")[i].checked=true
};\n
'''

		f.write(kkkk)

		#主次卧调换
		if int(returndata.rentfl)%100 == 0:
			f.write('document.getElementsByName("rentroom")[0].value="次卧"\n')
		else:
			f.write('document.getElementsByName("rentroom")[0].value="主卧"\n')

		#标题调换
		# title = (returndata.tier).encode("utf-8")+"环境优雅 闹中取静 交通便利 温馨舒适 阳光充足"
		# if int(returndata.rentfl)%100 == 0:
		# 	f.write('document.getElementById("address").value="%s"\n' % title)
		# elif int(returndata.rentfl)%50 == 0:
		# 	title = (returndata.tier).encode("utf-8")+"采光好 空间大 小区绿化好 配套齐全 环境舒适"
		# 	f.write('document.getElementById("address").value="%s"\n' % title)
		# else:
		# 	title = (returndata.tier).encode("utf-8")+"环境整洁清爽 采光格局好 冬暖夏凉 交通便利"
		# 	f.write('document.getElementById("address").value="%s"\n' % title)

		#每次更换标题信息
		if int(datanum) - othervarible.FIRSTNUM < 0:
			othervarible.FIRSTNUM = othervarible.FIRSTNUM - 3

			
		print othervarible.FIRSTNUM
		# if int(datanum)-othervarible.FIRSTNUM== 0:
		# 	title = (returndata.tier).encode("utf-8")+"环境优雅 闹中取静 交通便利 温馨舒适 阳光充足"
		# 	f.write('document.getElementById("address").value="%s"\n' % title)
		# elif int(datanum)-othervarible.FIRSTNUM== 1:
		# 	title = (returndata.tier).encode("utf-8")+"采光好 空间大 小区绿化好 配套齐全 环境舒适"
		# 	f.write('document.getElementById("address").value="%s"\n' % title)
		# elif int(datanum)-othervarible.FIRSTNUM== 2:
		# 	title = (returndata.tier).encode("utf-8")+"环境整洁清爽 采光格局好 冬暖夏凉 交通便利"
		# 	f.write('document.getElementById("address").value="%s"\n' % title)
		# 	othervarible.FIRSTNUM= othervarible.FIRSTNUM+3
		if int(datanum)-othervarible.FIRSTNUM== 0:
			title = (returndata.tier).encode("utf-8")+"精装现房，格局方正，男女不限，随时住"
			f.write('document.getElementById("address").value="%s"\n' % title)
		elif int(datanum)-othervarible.FIRSTNUM== 1:
			title = (returndata.tier).encode("utf-8")+"地铁沿线，南北通透，性价比高，采光好"
			f.write('document.getElementById("address").value="%s"\n' % title)
		elif int(datanum)-othervarible.FIRSTNUM== 2:
			title = (returndata.tier).encode("utf-8")+"干净整洁，正规成熟，随时看房，无中介"
			f.write('document.getElementById("address").value="%s"\n' % title)
			othervarible.FIRSTNUM= othervarible.FIRSTNUM+3


		selectpic = '''
setTimeout(function(){
for (var i = 1; i < 5; i++) {
document.getElementsByClassName("btn")[i].click()
}
},1000)
''' 
		f.write(selectpic)
		#拿取小区地址自动填入
# 		idjavascript = '''
# function myTimer() {
# var xmlhttp;
# xmlhttp = new XMLHttpRequest();
# xmlhttp.onreadystatechange = function() {
#     if (xmlhttp.readyState == 4 && xmlhttp.status == 200) {
        
#         console.log(JSON.parse(xmlhttp.responseText))
#         document.getElementById("district").value = JSON.parse(xmlhttp.responseText).district
#         document.getElementById("street").value = JSON.parse(xmlhttp.responseText).streetname
#         document.getElementById("blockaddress").value = JSON.parse(xmlhttp.responseText).address
#         document.getElementById("blockshowname").value=JSON.parse(xmlhttp.responseText).blockname
#         window.scroll(100,1650);
#     } 
# }
# xmlhttp.open("GET", "http://nj.zsb.house365.com/ajax/find_block_detail/?id=%s" , true);

# xmlhttp.send();

# };
# myTimer();\n

# document.onkeydown=function(event){
#   var e = event || window.event || arguments.callee.caller.arguments[0];
            
#    // if(e && e.keyCode==13){ // enter 键
#    if(e && e.keyCode==32){ // space 键
#        document.getElementById("jsBtnSubmit_rent").click();
#   }
# }; 

# 		''' % (str(getid((returndata.tier).encode("utf-8"))))

# 		f.write(idjavascript)
		
		# content = '''
		# 嗨住的优质房源
		# 【便利的交通&成熟的配套】
		# 超市，银行，医院，学校，便利店，菜市场，购物广场应有尽有,安全性高、位置安静、靠近花园、采光什么都非常好
		# 【装修精致\配套成熟的整租房源】

		# 厨房、卫生间、热水器、空调、WIFI,设备齐全，拎包就能入住,还有送车位的哦~
		# 【有趣善良的合租室友】
		# 那些直播里的萌妹纸、那些陆家嘴的金融GG、那些一身正气的程序员，可能就是你隔壁的室友~那些合租的小伙伴都是素质住户
		# 【品质小区】
		# 小区都处于繁华地段、房型正、佳、环境非常优美、都是一些次新小区
		# 嗨住的优质服务
		# * 无任何中介费
		# 嗨住房源实地勘察，保证嗨住租房平台房源可租信息实时可靠
		# * 房源广（7x24不断更新中）
		# 上海房源48396、北京16411、杭州23495、南京12055
		# * 直呼房东
		# 租客可通过嗨住租房上公开的房东电话，直接呼叫联系房东，询问房屋信息、预约看 房、达成交易。
		# * 免费预约
		# 租客可直接在嗨住租房平台提交对某套房的看房预约请求，嗨住顾问将免费为租客协调与房东的见面时间，与房东直接见面，达成交易。
		# * 租后服务
		# 嗨住还为租客提供了房租分期、室友征选、搬家维修等系列租后服务。
		# 更多房源欢迎致电4008180555(上海4008180555，杭州4008170019，南京4001606145）
		# 或
		# 至任意手机应用市场下载“嗨住租房”app


		# '''

		

	# return render(request,"home.html",{"tier":returntier})
	#解决AJAX跨域访问问题
	#No 'Access-Control-Allow-Origin' header
	response = HttpResponse(json.dumps(returntier))
	response["Access-Control-Allow-Origin"] = "*"
	response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
	response["Access-Control-Max-Age"] = "1000"
	response["Access-Control-Allow-Headers"] = "*"
	return response

def uploadpic(request):
	if request.method=='POST':
		photo=request.FILES['photo']
		if photo:
			img=Image.open(photo)
			img.save('D:\codes\\'+str(photo))
		return HttpResponse ("ok")
	obj = render(request,"upload.html")

	##放入cookie
	# obj.set_cookie("upload","syy")

	##检查cookie,如果cookie不正确，禁止登录
	# upload = request.COOKIES.get('upload')
	# # print  upload
	# if upload != "yy":
	# 	return HttpResponse ("None")

	##设置session
	# request.session['upload'] = "syy"

	##session判断登录
	# upload = request.session.get('upload')
	# if upload != "yy":
	# 	return HttpResponse ("None")

	##查询session的key,value
	from django.contrib.sessions.models import Session

	##session_key 这个可以从客户端浏览器或是服务器数据库
	session_key = '8u627zl56v35kky8exa3pplhruci01f2'

	session = Session.objects.get(session_key=session_key)
	print  session.get_decoded()

	return obj

def writeAnjukeDate(request,datanum):
	returndata = Person.objects.all()[int(datanum)]
	print int(datanum)
	print returndata.tier
	housenum = len(Person.objects.all())
	returntier = {"tier":(returndata.tier).encode("utf-8"),"price":returndata.rentfl.encode("utf-8"),"housenum":str(housenum)}


	readscriptpath =  os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))+"\\fillAnjuke\\readscript.js"

	with open(readscriptpath,"w+") as f:
		fillAnjuke = '''
			

			document.getElementsByName("room")[0].value = %s ;
			document.getElementsByName("hall")[0].value = %s ;
			document.getElementsByName("bathroom")[0].value = %s ;
			document.getElementsByName("roomarea")[0].value = %s ;
			document.getElementsByName("rentprice")[0].value = %s ;

			
			//房屋类型
			document.getElementsByName("housetype")[0].options[0].selected = false;
			document.getElementsByName("housetype")[0].options[1].selected =true;
			//装修情况
			document.getElementsByName("housefit")[0].options[0].selected = false;
			document.getElementsByName("housefit")[0].options[3].selected =true;
			//房间朝向
			document.getElementsByName("roomorient")[0].options[0].selected = false;
			document.getElementsByName("roomorient")[0].options[9].selected =true;
			//付款方式
			document.getElementsByName("paymode")[0].options[0].selected = false;
			document.getElementsByName("paymode")[0].options[1].selected =true;
			//全选
			for (var i = document.getElementsByName("fitment[]").length - 1; i >= 0; i--) {
				document.getElementsByName("fitment[]")[i].click();
			};
			
			//选择模板
			document.getElementsByClassName("use-tpl ui-button ui-button-blue ui-button-micro")[0].click();
			
			setTimeout(function(){
			document.getElementsByName("roomtemplate")[0].click();

			//模版确认延迟点击
			document.getElementsByClassName("ui-button ui-button-positive ui-button-medium")[0].click();
			
			},100)

			
			document.getElementsByName("community_unite")[0].value = '%s';

			function showHint() {
						    var xmlhttp;

						    if (window.XMLHttpRequest) {
						        // IE7+, Firefox, Chrome, Opera, Safari 浏览器执行代码
						        xmlhttp = new XMLHttpRequest();
						    } else {
						        // IE6, IE5 浏览器执行代码
						        xmlhttp = new ActiveXObject("Microsoft.XMLHTTP");
						    }
						    xmlhttp.onreadystatechange = function() {
						   			        if (xmlhttp.readyState == 4 && xmlhttp.status == 200) {
						   			            console.log(JSON.parse(xmlhttp.responseText)["data"][0]["name"]);
						   			            document.getElementsByName("communityAJK")[0].value = JSON.parse(xmlhttp.responseText)["data"][0]["name"];

						   			            	for (var i = document.getElementsByName("zoneAJK")[0].options.length - 1; i >= 0; i--) {
						   			            			                if (document.getElementsByName("zoneAJK")[0].options[i].text===JSON.parse(xmlhttp.responseText)["data"][0]["area_string"]){
						   			            			                    document.getElementsByName("zoneAJK")[0].options[0].selected = false;
						   			            			                    document.getElementsByName("zoneAJK")[0].options[i].selected = true;
						   			            			                    console.log(document.getElementsByName("zoneAJK")[0].options[i]);
						   			            			                    console.log(document.getElementsByName("zoneAJK")[0].value);
						   			            			                    showHint1(document.getElementsByName("zoneAJK")[0].value,JSON.parse(xmlhttp.responseText)["data"][0]["sub_area_string"]);
						   			            			                }
						   			            			            }

						   			           	// document.getElementsByClassName("ui-select-label")[0].innerText = JSON.parse(xmlhttp.responseText)["data"][0]["area_string"];
						   			           
						   			            // document.getElementsByClassName("ui-select-label")[1].innerText = JSON.parse(xmlhttp.responseText)["data"][0]["sub_area_string"];

						   			            document.getElementsByName("addressAJK")[0].value =JSON.parse(xmlhttp.responseText)["data"][0]["address"];

						   			            // document.getElementById("ajk_community_address").innerText = "地址："+ JSON.parse(xmlhttp.responseText)["data"][0]["area_string"]+" "
						   			            // +JSON.parse(xmlhttp.responseText)["data"][0]["sub_area_string"]+" "+JSON.parse(xmlhttp.responseText)["data"][0]["address"]

						   			        }
						   			    }
						    xmlhttp.open("GET", "http://vip.anjuke.com/ajax/community/search/?q=%s", true);
						    xmlhttp.send();
						    
						}

						function showHint1(e,f) {
						    var xmlhttp;

						    if (window.XMLHttpRequest) {
						        // IE7+, Firefox, Chrome, Opera, Safari 浏览器执行代码
						        xmlhttp = new XMLHttpRequest();
						    } else {
						        // IE6, IE5 浏览器执行代码
						        xmlhttp = new ActiveXObject("Microsoft.XMLHTTP");
						    }
						    xmlhttp.onreadystatechange = function() {
						   			        if (xmlhttp.readyState == 4 && xmlhttp.status == 200) {
						   			            for (var i = JSON.parse(xmlhttp.responseText)["data"].length - 1; i >= 0; i--) {
						   			                if (JSON.parse(xmlhttp.responseText)["data"][i]["typeName"]==f){
						   			                    // document.getElementsByName("blockAJK")[0].options[0].selected = false;
						   			                    // document.getElementsByName("blockAJK")[0].options[i].selected = true;
								   			            console.log(JSON.parse(xmlhttp.responseText)["data"][i]["typeName"])
								   			            console.log(JSON.parse(xmlhttp.responseText)["data"][i]["typeId"])
								   			            document.getElementsByName("blockAJK")[0].value = JSON.parse(xmlhttp.responseText)["data"][i]["typeId"]
								   			            // console.log(document.getElementsByName("blockAJK")[0].options[i])
						   			                }
						   			            }
						   			        }
						   			    }
						    xmlhttp.open("GET", "http://vip.anjuke.com/ajax/house/hz_house?act=getBlocks&districtId="+e, true);
						    xmlhttp.send();
						    
						}
						//showHint();



		''' % (returndata.houseinfo.encode("utf-8").split("室")[0],returndata.houseinfo.encode("utf-8").split("室")[1].split("厅")[0],
				returndata.houseinfo.encode("utf-8").split("室")[1].split("厅")[1][0],returndata.housecf.encode("utf-8").split('m²')[0],
				returndata.rentfl.encode("utf-8"),(returndata.tier).encode("utf-8"),(returndata.tier).encode("utf-8")
			)
		f.write(fillAnjuke)

		if int(returndata.houseinfo.encode("utf-8").split("室")[0]) > 3:
			# f.write('document.getElementsByName("flatshare")[0].value = 3;\n')
			f.write('document.getElementsByName("flatshare")[0].value = %s;\n' % returndata.houseinfo.encode("utf-8").split("室")[0])
		else:
			f.write('document.getElementsByName("flatshare")[0].value = %s;\n' % returndata.houseinfo.encode("utf-8").split("室")[0])

		if int(returndata.housecf5.encode("utf-8").split("/")[0])==0:
			f.write('document.getElementsByName("floor")[0].value = %s\n' % "3")
		else:
			f.write('document.getElementsByName("floor")[0].value = %s\n' % returndata.housecf5.encode("utf-8").split("/")[0])

		if int(returndata.housecf5.encode("utf-8").split("/")[1].split("层")[0])==0:
			f.write('document.getElementsByName("allFloor")[0].value = %s\n' % "7")
		else:
			f.write('document.getElementsByName("allFloor")[0].value = %s\n' % returndata.housecf5.encode("utf-8").split("/")[1].split("层")[0])

		#每次更换标题信息

		if int(datanum) == 0:##当刷新重新开始时,全部变量归零
			othervarible.FIRSTNUM = 0

		if int(datanum) - othervarible.FIRSTNUM < 0:
			othervarible.FIRSTNUM = othervarible.FIRSTNUM - 3
			
		print othervarible.FIRSTNUM
		if int(datanum)-othervarible.FIRSTNUM== 0:
			title = (returndata.tier).encode("utf-8")+"精装现房，格局方正，男女不限，随时住"
			f.write('document.getElementsByName("title")[0].value = "%s"\n' % title)
		elif int(datanum)-othervarible.FIRSTNUM== 1:
			title = (returndata.tier).encode("utf-8")+"南北通透，性价比高，采光好"	
			f.write('document.getElementsByName("title")[0].value = "%s"\n' % title)
		elif int(datanum)-othervarible.FIRSTNUM== 2:
			title = (returndata.tier).encode("utf-8")+"干净整洁，正规成熟，随时看房，无中介"
			f.write('document.getElementsByName("title")[0].value = "%s"\n' % title)
			othervarible.FIRSTNUM= othervarible.FIRSTNUM+3
		else:
			othervarible.FIRSTNUM = int(datanum)##当值超过三,永远使用为0的标题

		#//自动选择电梯
		allFloor = '''
		if (document.getElementsByName("allFloor")[0].value > 7) {
			document.getElementsByName("lift")[0].checked = true;
		}
		else{
			document.getElementsByName("lift")[1].checked = true;
		}
		'''
		f.write(allFloor+'\n')

		#过30秒自动点击发布
		# f.write('setTimeout(function(){ document.getElementById("publish-rent-add").click(); },20000);\n')

		#//自动选择安居库平台
		f.write('document.getElementById("chooseWeb_2").checked = true;\n')
		f.write('setTimeout(function(){ \
document.getElementsByClassName("ui-button ui-button-positive ui-button-medium")[2].click(); \
},100);\n' 
		)
		# f.write('document.getElementsByClassName("ui-button ui-button-positive ui-button-medium")[1].click();\n')

		#//刷新一次页面(未完成~~)
		freshPageOnce = '''  var isFirst = setTimeout(function(){history.go(0)},2000); 
   			 		 window.clearTimeout(isFirst);//去掉定时器 \n'''
		f.write(freshPageOnce)

		#//选择无中介费
		f.write('if (document.getElementsByName("noCommission")[0].checked ==false)\n{document.getElementsByName("noCommission")[0].click()};\n')

		#过30秒自动点击发布
		f.write('setTimeout(function(){ document.getElementById("publish-rent-add").click(); },30000);\n')


			
	response = HttpResponse(json.dumps(returntier))
	response["Access-Control-Allow-Origin"] = "*"
	response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
	response["Access-Control-Max-Age"] = "1000"
	response["Access-Control-Allow-Headers"] = "*"
	return response

def ftp(request):
	testselect = request.GET.get("test")
	print testselect
	from pyftpdlib.authorizers import DummyAuthorizer
	from pyftpdlib.handlers import FTPHandler
	from pyftpdlib.servers import FTPServer

	authorizer = DummyAuthorizer()

	authorizer.add_user('root', 'admin', '.', perm='elradfmw')

	authorizer.add_anonymous('.')

	handler = FTPHandler
	handler.authorizer = authorizer


	if testselect == "start":

		server = FTPServer(('10.137.7.124', 21), handler)
		server.serve_forever()
		# return HttpResponse(str(testselect))

	elif testselect == "stop":
		server = FTPServer(('10.137.7.124', 21), handler)
		server.close_all()
		# print "stop"
		print "now-stop"


	# print str(testselect)
	return render(request, 'ftp.html')

def get10086phone(request):
	phone = request.GET.get("a")
	print phone
	with open("phone.txt","a+") as phonetxt:
		phonetxt.write(phone+"\n")
	return HttpResponse (phone)


def getVip58Send(request):
	sending = request.GET.get("sending")
	stillSend = request.GET.get("stillSend")
	account = request.GET.get("account")
	with open("getvip58send.txt",'a+') as getvip58send:
		if account != '0':
			getvip58send.write(account+" ")
		elif stillSend != '0' :
			# print stillSend
			# print sending
			getvip58send.write(sending+" "+stillSend+" "+"\n")
	return HttpResponse(sending)

##嗨住后端接收数据接口
from openpyxl import load_workbook
import when

# FIRSTNAME = "徐汇"
FIRSTVAR = 0
VALUES = 0
ROW = 0
numList = []
def HZadmin(request):


	##发布套数
	HZalert = request.GET.get("HZalert")
	##发布区域
	# region1 = request.GET.get("region")

	# region = {"2":"静安","7":"卢湾","14":"黄浦","24":"徐汇","43":"长宁","53":"浦东","88":"虹口","97":"杨浦","108":"普陀","121":"闵行","140":"闸北","147":"宝山","166":"嘉定","177":"松江","197":"奉贤","206":"金山","215":"青浦","224":"崇明"}
	# print region[region1]
	# print HZalert

	# if region[region1] !=FIRSTNAME:
	# 	global FIRSTVAR
	# 	FIRSTVAR = 0
	# 	global FIRSTNAME
	# 	FIRSTNAME = region[region1]
	# 	global VALUES
	# 	VALUES = 0


	# ##装载表格
	# wb = load_workbook('58发房.xlsx')

	# ##确定当日时间
	# i = when.today()
	# today = str(i.month) + "月" + str(i.day) + "日"

	# ##如果当日sheet未建立的话，就建立
	# if today not in wb.sheetnames:##这里wb的语法有所改变
	# 	ws2 = wb.create_sheet(title=today)

	# ##选择今日sheet
	# sheet_ranges = wb[today]

	##按照excel列来生成字典
	excelNum = [3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,24,25,26,27,28,29,30,31,32,33]
	# region2 = [
	# "徐汇",
	# "长宁",
	# "浦东",
	# "浦东",
	# "杨浦",
	# "普陀",
	# "闵行",
	# "闵行",
	# "闸北",
	# "宝山",
	# "宝山",
	# "嘉定",
	# "嘉定",
	# "松江",
	# "松江",
	# "青浦",
	# ]
	# column = dict(zip(region2,excelNum))
	# print (column)

	allColumn = ["D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z",
	"AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ"]


	##写数据至表格(旧方法，已改进)
	# global FIRSTVAR
	# sheet_ranges[allColumn[FIRSTVAR]+str(excelNum[ROW])].value = int(HZalert)

	##准备改写，把所有数据保存至列表
	global numList
	b=numList.append(int(HZalert))
	print (numList)
	print (sum(numList))
	##统计列表和
	# VALUES = sum(numList)
	# print (VALUES)


	##所有列的和，统计是否超过100(原方法)
	# for grip in allColumn:
	# 	# print type(sheet_ranges[str(grip)+str(excelNum[ROW])].value)
	# 	if isinstance(sheet_ranges[str(grip)+str(excelNum[ROW])].value,int):
	# 		# print (sheet_ranges[str(grip)+str(excelNum[ROW])].value)
	# 		global VALUES
	# 		VALUES += sheet_ranges[str(grip)+str(excelNum[ROW])].value 
	# print (VALUES)

	##判断是否超过100，超过就下一区域(原方法，已改写)
	if sum(numList) >= 100:
		# print (numList)

		##装载表格
		wb = load_workbook('58发房.xlsx')

		##确定当日时间
		i = when.today()
		today = str(i.month) + "月" + str(i.day) + "日"

		##如果当日sheet未建立的话，就建立
		if today not in wb.sheetnames:##这里wb的语法有所改变
			ws2 = wb.create_sheet(title=today)

		##选择今日sheet
		sheet_ranges = wb[today]

		##循环把数据写入excel表格
		for i in numList:
			global FIRSTVAR
			sheet_ranges[allColumn[FIRSTVAR]+str(excelNum[ROW])].value = i
			##变量增加
			FIRSTVAR = FIRSTVAR + 1
		##保存表格
		wb.save('58发房.xlsx')

		global ROW
		ROW = ROW + 1
		FIRSTVAR = 0
		# VALUES = 0
		numList = []



	##变量增加(老方法，已不用)
	# FIRSTVAR = FIRSTVAR + 1
	# print FIRSTVAR




	response = HttpResponse(sum(numList))
	# response = HttpResponse(json.dumps({"values1":VALUES}))
	response["Access-Control-Allow-Origin"] = "*"
	response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
	response["Access-Control-Max-Age"] = "1000"
	response["Access-Control-Allow-Headers"] = "*"
	return response