# -*- coding: utf-8 -*-

from openpyxl import load_workbook

wb = load_workbook(u'58发房.xlsx')

##拿取所有行
for row in wb[u'6月26日'].iter_rows(min_col=5,min_row=3):
	k = 0
	for i in row:
		# print type(i.value)
		if isinstance(i.value,long):
			k = k + i.value
print k